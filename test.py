import openpyxl
from openpyxl import Workbook
import os

folder_name = "users_data.xlsx"

if not os.path.exists(folder_name):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Email", "Phone Number"])
    workbook.save(folder_name)

def add_user():
    name = input("Enter Your Name: ")
    email = input("Enter Your Email: ")
    phone = input("Enter Your Phone Number: ")

    workbook = openpyxl.load_workbook(folder_name)
    sheet = workbook.active
    sheet.append([name, email, phone])
    workbook.save(folder_name)
    print("Adding Your Details Successfully!\n")

def display_users():
    workbook = openpyxl.load_workbook(folder_name)
    sheet = workbook.active

    print("\nStored Users: ")
    for row in sheet.iter_rows(value_only = True):
        print(f"Name: {row[0]}, Email: {row[1]}, Phone Number: {row[2]}")
    print()

def main():
    while True:
        print("Choose an option: ")
        print("1. Add User")
        print("2. Display Users")
        print("3. Exit")

        choice = input("Please Enter Your Choice: ")

        if choice == '1':
            add_user()
        elif choice == '2':
            display_users()
        elif choice == '3':
            print("Existing program.")
            break
        else:
            print("Invalid Choice. Can You Please Try Again.\n")

if __name__ == "__main__":
    main()


